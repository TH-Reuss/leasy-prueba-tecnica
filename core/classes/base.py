from django.forms import ValidationError
from django.shortcuts import redirect
from django.views.generic import ListView
from django.utils.http import urlencode
from django.db.models import Q
from django.http import HttpResponse
import csv
import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.conf import settings
from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from openpyxl import Workbook


EXCEL_SUPPORTED = True

class CustomListView(ListView):
    available_columns = {}  # { 'Label': 'campo' }
    default_columns = []    # ['campo', ...]
    paginate_by = 20
    create_url = None
    update_url = None

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            self.object_list = self.get_queryset()
            export_type = request.GET.get('export', 'csv')
            return self.export_data(export_type)
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        columns_param = self.request.GET.get('columns')
        selected_attrs = columns_param.split(',') if columns_param else self.default_columns
        selected_attrs = selected_attrs if 'id' in selected_attrs else ['id'] + selected_attrs
        queryset = self.model.objects.values(*selected_attrs)

        search_query = self.request.GET.get('search')
        if search_query:
            filters = Q()
            for attr in selected_attrs:
                filters |= Q(**{f"{attr}__icontains": search_query})

            queryset = queryset.filter(filters)

        return queryset


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        valid_attr_values = set(self.available_columns.values())

        columns_param = self.request.GET.get('columns')
        if columns_param:
            selected_values = [
                val for val in columns_param.split(',') if val in valid_attr_values
            ]
        else:
            selected_values = self.default_columns

        selected_columns = {
            label: value for label, value in self.available_columns.items()
            if value in selected_values
        }

        query_params = self.request.GET.copy()
        query_params.pop('page', None)
        

        context.update({
            'available_columns': self.available_columns,
            'selected_columns': selected_columns,
            'query_string': urlencode(query_params, doseq=True),
            'create_url': self.create_url,
            'update_url': self.update_url
        })
        return context
    
    def export_data(self, export_type='csv'):
        queryset = self.get_queryset()
        context = self.get_context_data()
        selected_columns = context['selected_columns']  # dict: label → attr

        headers = list(selected_columns.keys())
        fields = list(selected_columns.values())

        if export_type == 'csv':
            response = HttpResponse(content_type='text/csv')
            filename = f"{self.model.__name__.lower()}s_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            writer = csv.writer(response)
            writer.writerow(headers)
            for row in queryset:
                writer.writerow([row.get(field, '') for field in fields])
            return response

        elif export_type == 'xlsx' and EXCEL_SUPPORTED:
            wb = Workbook()
            ws = wb.active
            ws.append(headers)

            for row in queryset:
                ws.append([str(row.get(field, '')) if row.get(field, '') is not None else '' for field in fields])

            for col_num, _ in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(col_num)].auto_size = True

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            filename = f"{self.model.__name__.lower()}s_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            wb.save(response)
            return response

        return HttpResponse(status=400)


class CustomCreateView(CreateView):
    template_name = 'core/form.html'
    title = 'Crear nuevo registro'

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, "Creado correctamente.")
        return response


    def get_success_url(self):
        return reverse_lazy(self.success_url)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = self.title
        return context


class CustomUpdateView(UpdateView):
    template_name = 'core/form.html'
    title = 'Editar registro'
    delete_url = None

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, "Actualizado correctamente.")
        return response
    
    def get_success_url(self):
        return reverse_lazy(self.success_url)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = self.title
        context['delete_url'] = self.delete_url
        return context
    

class CustomDeleteView(DeleteView):
    http_method_names = ['post']
    success_url = None

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            self.object.delete()
            messages.success(request, "Eliminado correctamente.")
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, "Error inesperado al intentar eliminar.")
        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy(self.success_url)
